from __future__ import annotations

import csv
import io
import json
import os
import re
import shutil
import sqlite3
from dataclasses import dataclass
from datetime import date, datetime
from functools import wraps
from pathlib import Path
from typing import Any

from flask import Flask, Response, flash, redirect, render_template, request, send_file, session, url_for
from openpyxl import load_workbook
from werkzeug.security import check_password_hash, generate_password_hash

from pos_pipeline import analyze_pos_exports, normalize_branch_name


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "instance"
UPLOAD_DIR = DATA_DIR / "uploads"
POS_RAW_DIR = DATA_DIR / "pos_raw"
POS_ANALYSIS_DIR = DATA_DIR / "pos_analysis"
DB_PATH = DATA_DIR / "pharmacy.db"
BRANCH_TEMPLATE_PATH = BASE_DIR / "pharmacy_branch_upload_template.xlsx"

ALLOWED_EXTENSIONS = {".xlsx", ".xlsm"}
ALLOWED_ANALYSIS_PERIOD_DAYS = (30, 90, 120)
DEAD_STOCK_MAX_SALES = 0
SLOW_MOVING_RATIO = 0.05
EXPIRY_WARNING_DAYS = 90
EXPIRY_CRITICAL_DAYS = 30
DEFAULT_ADMIN_USERNAME = os.environ.get("PHARMACY_ADMIN_USERNAME", "admin")
DEFAULT_ADMIN_PASSWORD = os.environ.get("PHARMACY_ADMIN_PASSWORD", "admin123")

ITEM_EXTRA_COLUMNS = {
    "category": "TEXT NOT NULL DEFAULT 'UNCATEGORIZED'",
    "cost_price": "REAL NOT NULL DEFAULT 0",
    "cost_value": "REAL NOT NULL DEFAULT 0",
    "avg_daily_sales": "REAL NOT NULL DEFAULT 0",
    "days_cover_forecast": "REAL",
    "stockout_probability_7d": "REAL NOT NULL DEFAULT 0",
    "stockout_probability_14d": "REAL NOT NULL DEFAULT 0",
    "stockout_probability_30d": "REAL NOT NULL DEFAULT 0",
    "reorder_point": "REAL NOT NULL DEFAULT 0",
    "recommended_reorder_qty": "REAL NOT NULL DEFAULT 0",
    "risk_score": "INTEGER NOT NULL DEFAULT 1",
    "decision_label": "TEXT NOT NULL DEFAULT 'Healthy'",
    "lost_sales_value": "REAL NOT NULL DEFAULT 0",
    "frozen_capital_value": "REAL NOT NULL DEFAULT 0",
    "unit_profit": "REAL NOT NULL DEFAULT 0",
    "gross_margin_percent": "REAL NOT NULL DEFAULT 0",
    "break_even_price": "REAL NOT NULL DEFAULT 0",
    "max_safe_discount_percent": "REAL NOT NULL DEFAULT 0",
    "suggested_offer_price": "REAL NOT NULL DEFAULT 0",
    "offer_discount_percent": "REAL NOT NULL DEFAULT 0",
    "offer_profit_per_unit": "REAL NOT NULL DEFAULT 0",
    "offer_profit_if_sold": "REAL NOT NULL DEFAULT 0",
    "signal_color": "TEXT NOT NULL DEFAULT 'green'",
}

UPLOAD_EXTRA_COLUMNS = {
    "analysis_period_days": "INTEGER NOT NULL DEFAULT 90",
}


app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("PHARMACY_SECRET_KEY", "change-this-secret-key")
app.config["MAX_CONTENT_LENGTH"] = 24 * 1024 * 1024
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0
app.jinja_env.auto_reload = True

POS_UPLOAD_FILES = {
    "product_activity": "PRODUCT ACTIVITY",
    "stock": "STOCK",
    "stock_cost": "STOCK COST",
}

WEEKLY_COMPARISON_METRICS = [
    ("sales", "Sales", "total_net_sales"),
    ("stock", "Stock value", "total_stock_value"),
    ("dead", "Dead stock", "dead_stock"),
    ("risk", "High risk", "high_risk_count"),
    ("frozen", "Frozen capital", "frozen_capital_value"),
    ("lost", "Lost sales", "lost_sales_value"),
]

TRANSLATIONS = {
    "en": {
        "app_name": "Predictive POS Cockpit",
        "app_description": "A decision cockpit for pharmacy stock, sales risk, and branch performance.",
        "language_toggle": "عربي",
        "dashboard": "Dashboard",
        "reorder_now": "Reorder Now",
        "dead_stock": "Dead Stock",
        "overstock_risk": "Overstock Risk",
        "fast_moving": "Fast Moving",
        "slow_moving": "Slow Moving",
        "financial_panel": "Financial Panel",
        "weekly_comparison": "Weekly Comparison",
        "branch_upload": "Branch Upload",
        "imports": "Imports",
        "data_management": "Data Management",
        "password": "Password",
        "logout": "Logout",
        "decision_queue": "Decision Queue",
        "analysis": "Analysis",
        "operations": "Operations",
        "admin": "Admin",
        "view_all": "View all",
        "history": "History",
        "command_center": "General command center",
        "branch_dashboard": "branch dashboard",
        "priority_actions": "Priority actions",
        "highest_risk": "Highest risk score",
        "urgent_reorder": "Urgent reorder candidates",
        "highest_frozen": "Highest frozen capital",
        "fast_winners": "Fast moving winners",
        "latest_import": "Latest import",
        "login": "Login",
        "username": "Username",
        "current_password": "Current password",
        "new_password": "New password",
        "confirm_new_password": "Confirm new password",
        "save_password": "Save password",
        "account": "Account",
        "change_password": "Change password",
        "branch_filter": "Branch filter",
        "all_branches": "All branches",
        "reset": "Reset",
        "upload_excel": "Upload Excel",
        "back_to_dashboard": "Back to dashboard",
        "back_to_reports": "Back to reports",
        "export_csv": "Export CSV",
        "select_branch": "Select branch",
        "risk": "Risk",
        "branch": "Branch",
        "item": "Item",
        "ref": "Ref",
        "sales": "Sales",
        "stock": "Stock",
        "decision": "Decision",
        "stock_value": "Stock value",
        "lost_sales": "Lost sales",
        "frozen_capital": "Frozen capital",
        "high_risk": "High risk",
        "current_week": "Current week",
        "previous_week": "Previous week",
        "uploads": "Uploads",
        "no_data": "No data.",
        "no_rows_report": "No rows in this report.",
    },
    "ar": {
        "app_name": "Predictive POS Cockpit",
        "app_description": "كابينة قرارات لمخزون الصيدليات، مخاطر المبيعات، وأداء الفروع.",
        "language_toggle": "English",
        "dashboard": "لوحة المتابعة",
        "reorder_now": "اشتري أو حوّل دلوقتي",
        "dead_stock": "مخزون راكد",
        "overstock_risk": "زيادة مخزون",
        "fast_moving": "أصناف سريعة",
        "slow_moving": "أصناف بطيئة",
        "financial_panel": "الماليات",
        "weekly_comparison": "مقارنة أسبوعية",
        "branch_upload": "رفع ملفات فرع",
        "imports": "سجل الرفع",
        "data_management": "إدارة البيانات",
        "password": "كلمة المرور",
        "logout": "خروج",
        "decision_queue": "قرارات مهمة",
        "analysis": "تحليلات",
        "operations": "تشغيل",
        "admin": "إدارة",
        "view_all": "شوف الكل",
        "history": "السجل",
        "command_center": "مركز المتابعة العام",
        "branch_dashboard": "لوحة الفرع",
        "priority_actions": "قرارات محتاجة حركة",
        "highest_risk": "أعلى درجات خطورة",
        "urgent_reorder": "أصناف محتاجة شراء أو تحويل",
        "highest_frozen": "أكبر فلوس متجمّدة",
        "fast_winners": "أصناف سريعة ومهمة",
        "latest_import": "آخر رفع",
        "login": "تسجيل الدخول",
        "username": "اسم المستخدم",
        "current_password": "كلمة المرور الحالية",
        "new_password": "كلمة المرور الجديدة",
        "confirm_new_password": "تأكيد كلمة المرور الجديدة",
        "save_password": "حفظ كلمة المرور",
        "account": "الحساب",
        "change_password": "تغيير كلمة المرور",
        "branch_filter": "فلتر الفرع",
        "all_branches": "كل الفروع",
        "reset": "رجّع الكل",
        "upload_excel": "ارفع إكسل",
        "back_to_dashboard": "ارجع للوحة",
        "back_to_reports": "ارجع للتقارير",
        "export_csv": "صدّر CSV",
        "select_branch": "اختار الفرع",
        "risk": "الخطورة",
        "branch": "الفرع",
        "item": "الصنف",
        "ref": "الكود",
        "sales": "المبيعات",
        "stock": "المخزون",
        "decision": "القرار",
        "stock_value": "قيمة المخزون",
        "lost_sales": "مبيعات ضايعة",
        "frozen_capital": "رأس مال متجمّد",
        "high_risk": "خطورة عالية",
        "current_week": "الأسبوع الحالي",
        "previous_week": "الأسبوع اللي فات",
        "uploads": "مرات الرفع",
        "no_data": "مفيش بيانات.",
        "no_rows_report": "مفيش صفوف في التقرير ده.",
    },
}


@dataclass
class ParsedRow:
    name: str
    sku: str
    sold: float
    stock: float
    price: float
    expiry: date | None

    @property
    def stock_value(self) -> float:
        return self.stock * self.price

    @property
    def turnover_ratio(self) -> float:
        return self.sold / self.stock if self.stock > 0 else 0


def get_db() -> sqlite3.Connection:
    DATA_DIR.mkdir(exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    with get_db() as db:
        db.executescript(
            """
            CREATE TABLE IF NOT EXISTS uploads (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                branch TEXT NOT NULL,
                week_date TEXT NOT NULL,
                filename TEXT NOT NULL,
                uploaded_at TEXT NOT NULL,
                analysis_period_days INTEGER NOT NULL DEFAULT 90,
                total_items INTEGER NOT NULL,
                dead_count INTEGER NOT NULL,
                expiry_count INTEGER NOT NULL,
                slow_count INTEGER NOT NULL,
                fast_count INTEGER NOT NULL
            );

            CREATE TABLE IF NOT EXISTS items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                upload_id INTEGER NOT NULL,
                branch TEXT NOT NULL,
                week_date TEXT NOT NULL,
                name TEXT NOT NULL,
                category TEXT NOT NULL DEFAULT 'UNCATEGORIZED',
                sku TEXT,
                sold REAL NOT NULL,
                stock REAL NOT NULL,
                price REAL NOT NULL,
                cost_price REAL NOT NULL DEFAULT 0,
                stock_value REAL NOT NULL,
                cost_value REAL NOT NULL DEFAULT 0,
                turnover_ratio REAL NOT NULL,
                expiry TEXT,
                days_left INTEGER,
                tags TEXT NOT NULL,
                urgency TEXT,
                action TEXT,
                reason TEXT,
                avg_daily_sales REAL NOT NULL DEFAULT 0,
                days_cover_forecast REAL,
                stockout_probability_7d REAL NOT NULL DEFAULT 0,
                stockout_probability_14d REAL NOT NULL DEFAULT 0,
                stockout_probability_30d REAL NOT NULL DEFAULT 0,
                reorder_point REAL NOT NULL DEFAULT 0,
                recommended_reorder_qty REAL NOT NULL DEFAULT 0,
                risk_score INTEGER NOT NULL DEFAULT 1,
                decision_label TEXT NOT NULL DEFAULT 'Healthy',
                lost_sales_value REAL NOT NULL DEFAULT 0,
                frozen_capital_value REAL NOT NULL DEFAULT 0,
                unit_profit REAL NOT NULL DEFAULT 0,
                gross_margin_percent REAL NOT NULL DEFAULT 0,
                break_even_price REAL NOT NULL DEFAULT 0,
                max_safe_discount_percent REAL NOT NULL DEFAULT 0,
                suggested_offer_price REAL NOT NULL DEFAULT 0,
                offer_discount_percent REAL NOT NULL DEFAULT 0,
                offer_profit_per_unit REAL NOT NULL DEFAULT 0,
                offer_profit_if_sold REAL NOT NULL DEFAULT 0,
                signal_color TEXT NOT NULL DEFAULT 'green',
                FOREIGN KEY(upload_id) REFERENCES uploads(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                created_at TEXT NOT NULL
            );
            """
        )
        existing_item_columns = {row["name"] for row in db.execute("PRAGMA table_info(items)").fetchall()}
        for column, definition in ITEM_EXTRA_COLUMNS.items():
            if column not in existing_item_columns:
                db.execute(f"ALTER TABLE items ADD COLUMN {column} {definition}")
        existing_upload_columns = {row["name"] for row in db.execute("PRAGMA table_info(uploads)").fetchall()}
        for column, definition in UPLOAD_EXTRA_COLUMNS.items():
            if column not in existing_upload_columns:
                db.execute(f"ALTER TABLE uploads ADD COLUMN {column} {definition}")
        user_count = db.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        if user_count == 0:
            db.execute(
                "INSERT INTO users (username, password_hash, created_at) VALUES (?, ?, ?)",
                (
                    DEFAULT_ADMIN_USERNAME,
                    generate_password_hash(DEFAULT_ADMIN_PASSWORD),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                ),
            )


def login_required(view):
    @wraps(view)
    def wrapped_view(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)

    return wrapped_view


def current_language() -> str:
    return session.get("lang", "en") if session.get("lang") in TRANSLATIONS else "en"


def translate(key: str) -> str:
    lang = current_language()
    return TRANSLATIONS.get(lang, TRANSLATIONS["en"]).get(key, TRANSLATIONS["en"].get(key, key))


@app.context_processor
def inject_language_helpers() -> dict[str, Any]:
    lang = current_language()
    return {
        "lang": lang,
        "text_dir": "rtl" if lang == "ar" else "ltr",
        "next_lang": "en" if lang == "ar" else "ar",
        "t": translate,
    }


@app.route("/language/<lang_code>", methods=["POST"])
def set_language(lang_code: str) -> Response:
    session["lang"] = "ar" if lang_code == "ar" else "en"
    return redirect(request.referrer or url_for("index"))


def normalize_header(value: Any) -> str:
    return str(value or "").strip().casefold()


def find_col(headers: list[str], candidates: list[str]) -> int:
    normalized = [h.casefold() for h in headers]
    for candidate in candidates:
        needle = candidate.casefold()
        for index, header in enumerate(normalized):
            if needle in header:
                return index
    return -1


def parse_num(value: Any) -> float:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^0-9.\-]", "", str(value))
    try:
        return float(cleaned) if cleaned else 0
    except ValueError:
        return 0


def parse_date(value: Any) -> date | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def parse_analysis_period_days(value: Any) -> int | None:
    try:
        period = int(str(value).strip())
    except (TypeError, ValueError, AttributeError):
        return None
    return period if period > 0 else None


def extract_branch_name(filename: str) -> str:
    stem = Path(filename).stem
    stem = re.sub(r"_?\d{4}-\d{2}-\d{2}", "", stem)
    return stem.replace("_", " ").strip() or "فرع غير محدد"


def read_excel_rows(path: Path) -> list[ParsedRow]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    headers = [normalize_header(h) for h in rows[0]]
    columns = {
        "name": find_col(headers, ["اسم الصنف", "item name", "product name", "الصنف", "name"]),
        "sku": find_col(headers, ["كود الصنف", "sku", "barcode", "كود", "رمز"]),
        "sold": find_col(headers, ["الكمية المباعة", "مباع", "sold", "qty sold", "sales"]),
        "stock": find_col(headers, ["الرصيد الحالي", "رصيد", "stock", "balance", "qty"]),
        "expiry": find_col(headers, ["تاريخ انتهاء الصلاحية", "انتهاء", "expiry", "exp date", "expiry date"]),
        "price": find_col(headers, ["سعر الوحدة", "سعر", "price", "unit price"]),
    }

    parsed: list[ParsedRow] = []
    for row in rows[1:]:
        name = str(row[columns["name"]] or "").strip() if columns["name"] >= 0 else ""
        sku = str(row[columns["sku"]] or "").strip() if columns["sku"] >= 0 else ""
        if not name and not sku:
            continue
        parsed.append(
            ParsedRow(
                name=name,
                sku=sku,
                sold=parse_num(row[columns["sold"]]) if columns["sold"] >= 0 else 0,
                stock=parse_num(row[columns["stock"]]) if columns["stock"] >= 0 else 0,
                price=parse_num(row[columns["price"]]) if columns["price"] >= 0 else 0,
                expiry=parse_date(row[columns["expiry"]]) if columns["expiry"] >= 0 else None,
            )
        )
    return parsed


def classify_rows(rows: list[ParsedRow]) -> list[dict[str, Any]]:
    ratios = [row.turnover_ratio for row in rows if row.stock > 0]
    avg_ratio = sum(ratios) / len(ratios) if ratios else 0
    today = date.today()
    analyzed: list[dict[str, Any]] = []

    for row in rows:
        tags: list[str] = []
        reason = ""
        urgency = ""
        action = ""
        days_left = None

        if row.stock > 0 and row.sold == DEAD_STOCK_MAX_SALES:
            tags.append("راكد")
            reason = "مبيعات = صفر هذا الأسبوع - يحتاج عرض ترويجي"
        elif row.stock > 0 and 0 < row.turnover_ratio < SLOW_MOVING_RATIO:
            tags.append("بطيء")
        elif row.turnover_ratio > avg_ratio * 2 and row.sold > 0:
            tags.append("سريع")

        if row.expiry:
            days_left = (row.expiry - today).days
            if 0 <= days_left <= EXPIRY_WARNING_DAYS:
                tags.append("قرب انتهاء")
                urgency = "خطر" if days_left <= EXPIRY_CRITICAL_DAYS else "تحذير"
                if days_left <= 14:
                    action = "عرض فوري أو إرجاع للمورد"
                elif days_left <= 30:
                    action = "Bundle أو تخفيض سريع"
                else:
                    action = "ضع في واجهة العرض"

        analyzed.append(
            {
                "name": row.name,
                "sku": row.sku,
                "sold": row.sold,
                "stock": row.stock,
                "price": row.price,
                "stock_value": row.stock_value,
                "turnover_ratio": row.turnover_ratio,
                "expiry": row.expiry.isoformat() if row.expiry else "",
                "days_left": days_left,
                "tags": " | ".join(tags) if tags else "طبيعي",
                "urgency": urgency,
                "action": action,
                "reason": reason,
            }
        )
    return analyzed


def save_upload(
    filename: str,
    branch: str,
    week_date: str,
    rows: list[dict[str, Any]],
    analysis_period_days: int = 90,
) -> int:
    counts = {
        "dead": sum("راكد" in r["tags"] for r in rows),
        "expiry": sum("قرب انتهاء" in r["tags"] for r in rows),
        "slow": sum("بطيء" in r["tags"] for r in rows),
        "fast": sum("سريع" in r["tags"] for r in rows),
    }
    with get_db() as db:
        cursor = db.execute(
            """
            INSERT INTO uploads
            (branch, week_date, filename, uploaded_at, analysis_period_days, total_items, dead_count, expiry_count, slow_count, fast_count)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                branch,
                week_date,
                filename,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                analysis_period_days,
                len(rows),
                counts["dead"],
                counts["expiry"],
                counts["slow"],
                counts["fast"],
            ),
        )
        upload_id = cursor.lastrowid
        db.executemany(
            """
            INSERT INTO items
            (upload_id, branch, week_date, name, sku, sold, stock, price, stock_value, turnover_ratio,
             expiry, days_left, tags, urgency, action, reason)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    upload_id,
                    branch,
                    week_date,
                    row["name"],
                    row["sku"],
                    row["sold"],
                    row["stock"],
                    row["price"],
                    row["stock_value"],
                    row["turnover_ratio"],
                    row["expiry"],
                    row["days_left"],
                    row["tags"],
                    row["urgency"],
                    row["action"],
                    row["reason"],
                )
                for row in rows
            ],
        )
    return int(upload_id)


def parse_optional_float(value: str | None) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except ValueError:
        return None


def parse_int(value: str | None, default: int = 0) -> int:
    try:
        return int(float(value or default))
    except ValueError:
        return default


def load_analysis_summary(path: Path) -> dict[str, Any]:
    summary_path = path / "summary.json"
    if not summary_path.exists():
        return {}
    return json.loads(summary_path.read_text(encoding="utf-8"))


TAG_ACTIONS = {
    "dead_stock": "Decision: discount, bundle, transfer, or return to supplier.",
    "stockout_with_sales": "Decision: urgent reorder or transfer stock from another branch.",
    "overstock_risk": "Decision: stop reordering and push with offer/transfer.",
    "fast_moving": "Decision: protect availability and raise reorder point.",
    "slow_moving": "Decision: review shelf placement and promo.",
}


def action_for_tags(tags: str) -> str:
    actions = [action for tag, action in TAG_ACTIONS.items() if tag in tags]
    return " ".join(actions)


def import_pos_analysis_to_db(path: Path, replace_all: bool = False) -> int:
    init_db()
    summary = load_analysis_summary(path)
    metadata = summary.get("metadata", {})
    branch = normalize_branch_name(metadata.get("branch") or path.name.split("_")[0])
    week_date = (metadata.get("end_date") or date.today().isoformat()).split(" ")[0]
    source_file = f"{path.name}/merged_items.csv"

    rows: list[dict[str, Any]] = []
    with (path / "merged_items.csv").open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            tags = row.get("tags") or "normal"
            row_branch = normalize_branch_name(row.get("branch") or branch)
            if row_branch == "i-unknown":
                row_branch = branch
            rows.append(
                {
                    "branch": row_branch,
                    "week_date": week_date,
                    "name": row.get("name") or "",
                    "category": row.get("category") or "UNCATEGORIZED",
                    "sku": row.get("ref") or "",
                    "sold": parse_num(row.get("pos_sales")),
                    "stock": parse_num(row.get("qoh")),
                    "price": parse_num(row.get("sales_price")),
                    "cost_price": parse_num(row.get("cost")),
                    "stock_value": parse_num(row.get("stock_value")),
                    "cost_value": parse_num(row.get("cost_value")),
                    "turnover_ratio": parse_num(row.get("turnover_ratio")),
                    "avg_daily_sales": parse_num(row.get("avg_daily_sales")),
                    "days_cover_forecast": parse_optional_float(row.get("days_cover")),
                    "stockout_probability_7d": parse_num(row.get("stockout_probability_7d")),
                    "stockout_probability_14d": parse_num(row.get("stockout_probability_14d")),
                    "stockout_probability_30d": parse_num(row.get("stockout_probability_30d")),
                    "reorder_point": parse_num(row.get("reorder_point")),
                    "recommended_reorder_qty": parse_num(row.get("recommended_reorder_qty")),
                    "risk_score": parse_int(row.get("risk_score"), default=1),
                    "decision_label": row.get("decision_label") or "Healthy",
                    "lost_sales_value": parse_num(row.get("lost_sales_value")),
                    "frozen_capital_value": parse_num(row.get("frozen_capital_value")),
                    "unit_profit": parse_num(row.get("unit_profit")),
                    "gross_margin_percent": parse_num(row.get("gross_margin_percent")),
                    "break_even_price": parse_num(row.get("break_even_price")),
                    "max_safe_discount_percent": parse_num(row.get("max_safe_discount_percent")),
                    "suggested_offer_price": parse_num(row.get("suggested_offer_price")),
                    "offer_discount_percent": parse_num(row.get("offer_discount_percent")),
                    "offer_profit_per_unit": parse_num(row.get("offer_profit_per_unit")),
                    "offer_profit_if_sold": parse_num(row.get("offer_profit_if_sold")),
                    "signal_color": row.get("signal_color") or "green",
                    "expiry": "",
                    "days_left": None,
                    "tags": tags,
                    "urgency": "High" if any(tag in tags for tag in ["stockout_with_sales", "dead_stock"]) else "",
                    "action": action_for_tags(tags),
                    "reason": "",
                }
            )

    counts = {
        "dead": sum("dead_stock" in row["tags"] for row in rows),
        "stockout": sum("stockout_with_sales" in row["tags"] for row in rows),
        "overstock": sum("overstock_risk" in row["tags"] for row in rows),
        "slow": sum("slow_moving" in row["tags"] for row in rows),
        "fast": sum("fast_moving" in row["tags"] for row in rows),
    }

    with get_db() as db:
        if replace_all:
            db.execute("DELETE FROM items")
            db.execute("DELETE FROM uploads")
        else:
            db.execute("DELETE FROM items WHERE branch = ?", (branch,))
            db.execute("DELETE FROM uploads WHERE branch = ?", (branch,))

        cursor = db.execute(
            """
            INSERT INTO uploads
            (branch, week_date, filename, uploaded_at, analysis_period_days, total_items, dead_count, expiry_count, slow_count, fast_count)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                branch,
                week_date,
                source_file,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                parse_int(metadata.get("period_days"), default=90),
                len(rows),
                counts["dead"],
                counts["stockout"] + counts["overstock"],
                counts["slow"],
                counts["fast"],
            ),
        )
        upload_id = cursor.lastrowid
        db.executemany(
            """
            INSERT INTO items
            (upload_id, branch, week_date, name, category, sku, sold, stock, price, cost_price, stock_value, cost_value, turnover_ratio,
             expiry, days_left, tags, urgency, action, reason, avg_daily_sales, days_cover_forecast,
             stockout_probability_7d, stockout_probability_14d, stockout_probability_30d, reorder_point,
             recommended_reorder_qty, risk_score, decision_label, lost_sales_value, frozen_capital_value,
             unit_profit, gross_margin_percent, break_even_price, max_safe_discount_percent,
             suggested_offer_price, offer_discount_percent, offer_profit_per_unit, offer_profit_if_sold, signal_color)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    upload_id,
                    row["branch"],
                    row["week_date"],
                    row["name"],
                    row["category"],
                    row["sku"],
                    row["sold"],
                    row["stock"],
                    row["price"],
                    row["cost_price"],
                    row["stock_value"],
                    row["cost_value"],
                    row["turnover_ratio"],
                    row["expiry"],
                    row["days_left"],
                    row["tags"],
                    row["urgency"],
                    row["action"],
                    row["reason"],
                    row["avg_daily_sales"],
                    row["days_cover_forecast"],
                    row["stockout_probability_7d"],
                    row["stockout_probability_14d"],
                    row["stockout_probability_30d"],
                    row["reorder_point"],
                    row["recommended_reorder_qty"],
                    row["risk_score"],
                    row["decision_label"],
                    row["lost_sales_value"],
                    row["frozen_capital_value"],
                    row["unit_profit"],
                    row["gross_margin_percent"],
                    row["break_even_price"],
                    row["max_safe_discount_percent"],
                    row["suggested_offer_price"],
                    row["offer_discount_percent"],
                    row["offer_profit_per_unit"],
                    row["offer_profit_if_sold"],
                    row["signal_color"],
                )
                for row in rows
            ],
        )
    return len(rows)


def branch_options() -> list[str]:
    with get_db() as db:
        return [row["branch"] for row in db.execute("SELECT DISTINCT branch FROM items ORDER BY branch").fetchall()]


def branch_data_summary() -> list[dict[str, Any]]:
    with get_db() as db:
        rows = db.execute(
            """
            SELECT
                branch,
                COUNT(*) AS item_count,
                COALESCE(SUM(stock_value), 0) AS stock_value,
                COALESCE(SUM(lost_sales_value), 0) AS lost_sales_value,
                MAX(week_date) AS latest_week
            FROM items
            GROUP BY branch
            ORDER BY branch
            """
        ).fetchall()
        upload_counts = {
            row["branch"]: row["upload_count"]
            for row in db.execute(
                """
                SELECT branch, COUNT(*) AS upload_count
                FROM uploads
                GROUP BY branch
                """
            ).fetchall()
        }

    summaries = []
    seen = set()
    for row in rows:
        branch = row["branch"]
        seen.add(branch)
        summaries.append(
            {
                "branch": branch,
                "item_count": row["item_count"],
                "upload_count": upload_counts.get(branch, 0),
                "stock_value": row["stock_value"],
                "lost_sales_value": row["lost_sales_value"],
                "latest_week": row["latest_week"] or "-",
            }
        )

    for branch, upload_count in sorted(upload_counts.items()):
        if branch not in seen:
            summaries.append(
                {
                    "branch": branch,
                    "item_count": 0,
                    "upload_count": upload_count,
                    "stock_value": 0,
                    "lost_sales_value": 0,
                    "latest_week": "-",
                }
            )

    return summaries


def remove_branch_files(branch: str | None = None) -> int:
    removed = 0
    targets: list[Path] = []
    if branch:
        targets.append(POS_RAW_DIR / branch)
        if POS_ANALYSIS_DIR.exists():
            targets.extend(POS_ANALYSIS_DIR.glob(f"{branch}_*"))
    else:
        targets.extend([POS_RAW_DIR, POS_ANALYSIS_DIR, UPLOAD_DIR])

    for target in targets:
        try:
            resolved = target.resolve()
        except FileNotFoundError:
            continue
        if not resolved.exists() or DATA_DIR.resolve() not in resolved.parents:
            continue
        if resolved.is_dir():
            shutil.rmtree(resolved)
            removed += 1
        else:
            resolved.unlink()
            removed += 1

    for directory in (UPLOAD_DIR, POS_RAW_DIR, POS_ANALYSIS_DIR):
        directory.mkdir(parents=True, exist_ok=True)

    return removed


def delete_branch_data(branch: str | None = None, delete_files: bool = True) -> dict[str, int]:
    with get_db() as db:
        if branch:
            item_count = db.execute("SELECT COUNT(*) FROM items WHERE branch = ?", (branch,)).fetchone()[0]
            upload_count = db.execute("SELECT COUNT(*) FROM uploads WHERE branch = ?", (branch,)).fetchone()[0]
            db.execute("DELETE FROM items WHERE branch = ?", (branch,))
            db.execute("DELETE FROM uploads WHERE branch = ?", (branch,))
        else:
            item_count = db.execute("SELECT COUNT(*) FROM items").fetchone()[0]
            upload_count = db.execute("SELECT COUNT(*) FROM uploads").fetchone()[0]
            db.execute("DELETE FROM items")
            db.execute("DELETE FROM uploads")

    file_count = remove_branch_files(branch) if delete_files else 0
    return {"items": item_count, "uploads": upload_count, "files": file_count}


def selected_branch(value: str | None) -> str | None:
    if not value or value == "all":
        return None
    branch = normalize_branch_name(value)
    return branch if re.fullmatch(r"i-(?:[1-9]|10)", branch) else None


def branch_query_args(branch: str | None) -> dict[str, str]:
    return {"branch": branch} if branch else {}


def fetch_items(category: str = "all", limit: int | None = None, branch: str | None = None) -> list[sqlite3.Row]:
    conditions: list[str] = []
    order = "id DESC"
    params: list[Any] = []
    if category == "dead":
        conditions.append("tags LIKE ?")
        params.append("%dead_stock%")
        order = "frozen_capital_value DESC, stock_value DESC"
    elif category == "expiry":
        conditions.append("tags LIKE ?")
        params.append("%stockout_with_sales%")
        order = "risk_score DESC, lost_sales_value DESC, sold DESC"
    elif category == "overstock":
        conditions.append("tags LIKE ?")
        params.append("%overstock_risk%")
        order = "risk_score DESC, frozen_capital_value DESC, stock_value DESC"
    elif category == "slow":
        conditions.append("tags LIKE ?")
        params.append("%slow_moving%")
        order = "cost_value DESC, days_cover_forecast DESC, stock_value DESC"
    elif category == "fast":
        conditions.append("tags LIKE ?")
        params.append("%fast_moving%")
        order = "sold DESC"
    elif category == "risk":
        conditions.append("risk_score >= ?")
        params.append(8)
        order = "risk_score DESC, lost_sales_value DESC, frozen_capital_value DESC"

    if branch:
        conditions.append("branch = ?")
        params.append(branch)

    where = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    sql = f"SELECT * FROM items {where} ORDER BY {order}"
    if limit:
        sql += " LIMIT ?"
        params.append(limit)
    with get_db() as db:
        return db.execute(sql, params).fetchall()


def dashboard_stats(branch: str | None = None) -> dict[str, Any]:
    where = "WHERE branch = ?" if branch else ""
    params = [branch] if branch else []
    with get_db() as db:
        stats = db.execute(
            f"""
            SELECT
                COUNT(*) AS total_items,
                SUM(CASE WHEN tags LIKE '%dead_stock%' THEN 1 ELSE 0 END) AS dead_count,
                SUM(CASE WHEN tags LIKE '%stockout_with_sales%' THEN 1 ELSE 0 END) AS expiry_count,
                SUM(CASE WHEN tags LIKE '%slow_moving%' THEN 1 ELSE 0 END) AS slow_count,
                SUM(CASE WHEN tags LIKE '%fast_moving%' THEN 1 ELSE 0 END) AS fast_count,
                SUM(CASE WHEN tags LIKE '%overstock_risk%' THEN 1 ELSE 0 END) AS overstock_count,
                SUM(CASE WHEN tags LIKE '%dead_stock%' THEN stock_value ELSE 0 END) AS dead_value,
                SUM(CASE WHEN tags LIKE '%overstock_risk%' THEN stock_value ELSE 0 END) AS overstock_value,
                SUM(lost_sales_value) AS lost_sales_value,
                SUM(frozen_capital_value) AS frozen_capital_value,
                SUM(recommended_reorder_qty) AS reorder_qty_total,
                SUM(CASE WHEN risk_score >= 8 THEN 1 ELSE 0 END) AS high_risk_count,
                AVG(risk_score) AS avg_risk_score,
                COUNT(DISTINCT branch) AS branch_count
            FROM items
            {where}
            """,
            params,
        ).fetchone()
        if branch:
            uploads = db.execute("SELECT * FROM uploads WHERE branch = ? ORDER BY id DESC LIMIT 8", (branch,)).fetchall()
        else:
            uploads = db.execute("SELECT * FROM uploads ORDER BY id DESC LIMIT 8").fetchall()
    return {"stats": stats, "uploads": uploads}


@app.route("/")
@login_required
def index() -> str:
    branch = selected_branch(request.args.get("branch"))
    data = dashboard_stats(branch)
    return render_template(
        "dashboard.html",
        **data,
        selected_branch=branch,
        branch_options=branch_options(),
        branch_args=branch_query_args(branch),
        top_dead=fetch_items("dead", 5, branch),
        top_expiry=fetch_items("expiry", 5, branch),
        top_overstock=fetch_items("overstock", 5, branch),
        top_fast=fetch_items("fast", 5, branch),
        top_risk=fetch_items("risk", 5, branch),
    )


@app.route("/upload", methods=["GET", "POST"])
@login_required
def upload() -> str | Response:
    if request.method == "GET":
        return render_template("upload.html", analysis_period_options=ALLOWED_ANALYSIS_PERIOD_DAYS)

    branch = normalize_branch_name(request.form.get("branch", ""))
    if not re.fullmatch(r"i-(?:[1-9]|10)", branch):
        flash("Choose a branch from i-1 to i-10.", "error")
        return redirect(url_for("upload"))

    analysis_period_mode = str(request.form.get("analysis_period_mode", "")).strip()
    analysis_period_days = (
        parse_analysis_period_days(request.form.get("analysis_period_days_custom"))
        if analysis_period_mode == "custom"
        else parse_analysis_period_days(analysis_period_mode)
    )
    if analysis_period_days is None:
        flash("Choose 30, 90, 120, or enter a custom positive number of days.", "error")
        return redirect(url_for("upload"))

    files = {key: request.files.get(key) for key in POS_UPLOAD_FILES}
    missing = [label for key, label in POS_UPLOAD_FILES.items() if not files[key] or not files[key].filename]
    if missing:
        flash(f"Upload all required POS files: {', '.join(missing)}.", "error")
        return redirect(url_for("upload"))

    for file in files.values():
        extension = Path(file.filename).suffix.lower()
        if extension not in ALLOWED_EXTENSIONS:
            flash("Supported files are .xlsx or .xlsm only.", "error")
            return redirect(url_for("upload"))

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    raw_dir = POS_RAW_DIR / branch / stamp
    raw_dir.mkdir(parents=True, exist_ok=True)
    saved_files: list[str] = []
    for key, label in POS_UPLOAD_FILES.items():
        file = files[key]
        assert file is not None
        extension = Path(file.filename).suffix.lower()
        saved_name = f"{branch}-{label}{extension}"
        file.save(raw_dir / saved_name)
        saved_files.append(saved_name)

    try:
        analysis_dir = analyze_pos_exports(
            raw_dir,
            POS_ANALYSIS_DIR,
            prefix=f"{branch}-",
            branch_override=branch,
            movement_period_days_override=analysis_period_days,
        )
        summary_path = analysis_dir / "summary.json"
        summary = json.loads(summary_path.read_text(encoding="utf-8"))
        metadata = summary.get("metadata", {})
        actual_period_days = int(metadata.get("actual_export_period_days") or 0)
        if actual_period_days != analysis_period_days:
            raise ValueError(
                f"Product Activity export covers {actual_period_days} days, but you selected {analysis_period_days} days."
            )
        imported_count = import_pos_analysis_to_db(analysis_dir, replace_all=False)
    except Exception as exc:
        flash(f"Could not analyze POS files: {exc}", "error")
        return redirect(url_for("upload"))

    flash(
        f"Imported {imported_count} POS rows for {branch} using a {analysis_period_days}-day analysis window. Raw files archived in {raw_dir}.",
        "success",
    )
    return redirect(url_for("index"))


@app.route("/download-template")
@login_required
def download_template() -> Response:
    if not BRANCH_TEMPLATE_PATH.exists():
        from create_branch_template import build_workbook

        build_workbook().save(BRANCH_TEMPLATE_PATH)
    return send_file(BRANCH_TEMPLATE_PATH, as_attachment=True, download_name="pharmacy_branch_upload_template.xlsx")


@app.route("/reports/<category>")
@login_required
def report(category: str) -> str:
    labels = {
        "all": "All Items",
        "dead": "Dead Stock",
        "expiry": "Stockout With Sales",
        "overstock": "Overstock Risk",
        "slow": "Slow Moving",
        "fast": "Fast Moving",
        "risk": "High Risk Score",
    }
    if category not in labels:
        return redirect(url_for("index"))
    branch = selected_branch(request.args.get("branch"))
    title = labels[category] if not branch else f"{labels[category]} - {branch}"
    return render_template(
        "report.html",
        title=title,
        category=category,
        rows=fetch_items(category, branch=branch),
        selected_branch=branch,
        branch_options=branch_options(),
        branch_args=branch_query_args(branch),
    )


@app.route("/branches/<branch>")
@login_required
def branch_dashboard(branch: str) -> Response:
    normalized = selected_branch(branch)
    if not normalized:
        return redirect(url_for("index"))
    return redirect(url_for("index", branch=normalized))


def financial_rows(branch: str | None = None) -> dict[str, Any]:
    item_where = "WHERE branch = ?" if branch else ""
    item_params = [branch] if branch else []
    with get_db() as db:
        branch_rows = db.execute(
            """
            SELECT
                branch,
                COUNT(*) AS items,
                SUM(lost_sales_value) AS lost_sales_value,
                SUM(frozen_capital_value) AS frozen_capital_value,
                SUM(CASE WHEN tags LIKE '%overstock_risk%' THEN stock_value ELSE 0 END) AS overstock_value,
                SUM(CASE WHEN tags LIKE '%dead_stock%' THEN stock_value ELSE 0 END) AS dead_value,
                SUM(CASE WHEN risk_score >= 8 THEN 1 ELSE 0 END) AS high_risk_count
            FROM items
            GROUP BY branch
            ORDER BY lost_sales_value DESC, frozen_capital_value DESC
            """
        ).fetchall()
        category_rows = db.execute(
            f"""
            SELECT
                category,
                COUNT(*) AS items,
                SUM(lost_sales_value) AS lost_sales_value,
                SUM(frozen_capital_value) AS frozen_capital_value,
                SUM(stock_value) AS stock_value,
                SUM(CASE WHEN risk_score >= 8 THEN 1 ELSE 0 END) AS high_risk_count
            FROM items
            {item_where}
            GROUP BY category
            ORDER BY lost_sales_value DESC, frozen_capital_value DESC
            LIMIT 12
            """,
            item_params,
        ).fetchall()
        totals = db.execute(
            f"""
            SELECT
                SUM(lost_sales_value) AS lost_sales_value,
                SUM(frozen_capital_value) AS frozen_capital_value,
                SUM(CASE WHEN tags LIKE '%overstock_risk%' THEN stock_value ELSE 0 END) AS overstock_value,
                SUM(CASE WHEN tags LIKE '%dead_stock%' THEN stock_value ELSE 0 END) AS dead_value,
                SUM(CASE WHEN risk_score >= 8 THEN 1 ELSE 0 END) AS high_risk_count
            FROM items
            {item_where}
            """,
            item_params,
        ).fetchone()

    branch_chart = [
        {
            "branch": row["branch"],
            "lostSales": round(float(row["lost_sales_value"] or 0), 2),
            "frozenCapital": round(float(row["frozen_capital_value"] or 0), 2),
            "overstock": round(float(row["overstock_value"] or 0), 2),
            "deadCapital": round(float(row["dead_value"] or 0), 2),
            "highRisk": int(row["high_risk_count"] or 0),
            "exposure": round(float((row["lost_sales_value"] or 0) + (row["frozen_capital_value"] or 0)), 2),
        }
        for row in branch_rows
    ]
    category_chart = [
        {
            "category": row["category"],
            "lostSales": round(float(row["lost_sales_value"] or 0), 2),
            "frozenCapital": round(float(row["frozen_capital_value"] or 0), 2),
            "stockValue": round(float(row["stock_value"] or 0), 2),
            "highRisk": int(row["high_risk_count"] or 0),
            "exposure": round(float((row["lost_sales_value"] or 0) + (row["frozen_capital_value"] or 0)), 2),
        }
        for row in category_rows
    ]
    trend_chart = financial_history(branch)
    max_branch_exposure = max([row["exposure"] for row in branch_chart] or [1])
    max_category_exposure = max([row["exposure"] for row in category_chart] or [1])
    return {
        "branch_rows": branch_rows,
        "category_rows": category_rows,
        "totals": totals,
        "max_branch_exposure": max_branch_exposure,
        "max_category_exposure": max_category_exposure,
        "branch_chart": branch_chart,
        "category_chart": category_chart,
        "trend_chart": trend_chart,
    }


def financial_history(branch: str | None = None) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if not POS_ANALYSIS_DIR.exists():
        return rows

    for summary_path in sorted(POS_ANALYSIS_DIR.glob("i-*_20*/summary.json"), key=lambda path: path.stat().st_mtime):
        try:
            summary = json.loads(summary_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        metadata = summary.get("metadata", {})
        item_branch = normalize_branch_name(metadata.get("branch") or summary_path.parent.name.split("_")[0])
        if branch and item_branch != branch:
            continue
        timestamp = summary_path.parent.name.split("_", 1)[-1]
        rows.append(
            {
                "label": timestamp[-6:-4] + ":" + timestamp[-4:-2] if len(timestamp) >= 6 else summary_path.parent.name,
                "branch": item_branch,
                "lostSales": round(float(summary.get("lost_sales_value") or 0), 2),
                "frozenCapital": round(float(summary.get("frozen_capital_value") or 0), 2),
                "highRisk": int(summary.get("high_risk_count") or 0),
                "reorderQty": round(float(summary.get("reorder_qty_total") or 0), 2),
                "exposure": round(float(summary.get("lost_sales_value") or 0) + float(summary.get("frozen_capital_value") or 0), 2),
            }
        )
    return rows[-18:]


def summary_sort_value(summary_path: Path, metadata: dict[str, Any]) -> str:
    end_date = str(metadata.get("end_date") or "").strip()
    if end_date:
        return end_date
    timestamp = summary_path.parent.name.split("_", 1)[-1]
    return timestamp or datetime.fromtimestamp(summary_path.stat().st_mtime).isoformat()


def summary_label(summary_path: Path, metadata: dict[str, Any]) -> str:
    start_date = str(metadata.get("start_date") or "").strip()
    end_date = str(metadata.get("end_date") or "").strip()
    if start_date and end_date:
        return f"{start_date} to {end_date}"
    if end_date:
        return end_date
    return summary_path.parent.name.split("_", 1)[-1]


def summary_metric(summary: dict[str, Any], metric_key: str) -> float:
    if metric_key == "dead_stock":
        return float((summary.get("tag_counts") or {}).get("dead_stock") or 0)
    return float(summary.get(metric_key) or 0)


def change_percent(current: float, previous: float) -> float | None:
    if previous == 0:
        return 0 if current == 0 else None
    return ((current - previous) / abs(previous)) * 100


def change_class(delta: float) -> str:
    if delta > 0:
        return "up"
    if delta < 0:
        return "down"
    return "flat"


def comparison_metric(current: float, previous: float | None) -> dict[str, Any]:
    previous_value = previous if previous is not None else 0
    delta = current - previous_value if previous is not None else 0
    percent = change_percent(current, previous_value) if previous is not None else None
    return {
        "current": round(current, 2),
        "previous": round(previous_value, 2),
        "delta": round(delta, 2),
        "percent": None if percent is None else round(percent, 1),
        "class": change_class(delta) if previous is not None else "flat",
    }


def weekly_summary_records(branch: str | None = None) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    if not POS_ANALYSIS_DIR.exists():
        return records

    for summary_path in POS_ANALYSIS_DIR.glob("i-*_20*/summary.json"):
        try:
            summary = json.loads(summary_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        metadata = summary.get("metadata", {})
        item_branch = normalize_branch_name(metadata.get("branch") or summary_path.parent.name.split("_")[0])
        if branch and item_branch != branch:
            continue
        records.append(
            {
                "branch": item_branch,
                "folder": summary_path.parent.name,
                "label": summary_label(summary_path, metadata),
                "sort_value": summary_sort_value(summary_path, metadata),
                "summary": summary,
            }
        )
    return sorted(records, key=lambda row: (row["branch"], row["sort_value"], row["folder"]))


def branch_weekly_comparison(branch: str | None = None) -> dict[str, Any]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for record in weekly_summary_records(branch):
        grouped.setdefault(record["branch"], []).append(record)

    rows: list[dict[str, Any]] = []
    for item_branch, records in sorted(grouped.items()):
        current = records[-1]
        previous = records[-2] if len(records) > 1 else None
        metrics = {}
        for key, _label, summary_key in WEEKLY_COMPARISON_METRICS:
            current_value = summary_metric(current["summary"], summary_key)
            previous_value = summary_metric(previous["summary"], summary_key) if previous else None
            metrics[key] = comparison_metric(current_value, previous_value)
        rows.append(
            {
                "branch": item_branch,
                "current_label": current["label"],
                "current_folder": current["folder"],
                "previous_label": previous["label"] if previous else "No previous upload",
                "previous_folder": previous["folder"] if previous else "",
                "upload_count": len(records),
                "has_previous": previous is not None,
                "metrics": metrics,
            }
        )

    totals = {}
    for key, _label, _summary_key in WEEKLY_COMPARISON_METRICS:
        current_total = sum(row["metrics"][key]["current"] for row in rows)
        comparable = [row for row in rows if row["has_previous"]]
        previous_total = sum(row["metrics"][key]["previous"] for row in comparable)
        totals[key] = comparison_metric(current_total, previous_total if comparable else None)

    return {"rows": rows, "totals": totals}


def product_history(sku: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if not POS_ANALYSIS_DIR.exists():
        return rows

    for csv_path in sorted(POS_ANALYSIS_DIR.glob("i-*_20*/merged_items.csv"), key=lambda path: path.stat().st_mtime):
        summary = load_analysis_summary(csv_path.parent)
        metadata = summary.get("metadata", {})
        folder_branch = normalize_branch_name(metadata.get("branch") or csv_path.parent.name.split("_")[0])
        label = metadata.get("end_date") or csv_path.parent.name.split("_", 1)[-1]
        try:
            with csv_path.open(encoding="utf-8-sig", newline="") as handle:
                for row in csv.DictReader(handle):
                    if str(row.get("ref") or "").strip() != sku:
                        continue
                    item_branch = normalize_branch_name(row.get("branch") or folder_branch)
                    if item_branch == "i-unknown":
                        item_branch = folder_branch
                    rows.append(
                        {
                            "label": label,
                            "folder": csv_path.parent.name,
                            "branch": item_branch,
                            "name": row.get("name") or "",
                            "category": row.get("category") or "UNCATEGORIZED",
                            "sold": parse_num(row.get("pos_sales")),
                            "stock": parse_num(row.get("qoh")),
                            "sales_price": parse_num(row.get("sales_price")),
                            "cost_price": parse_num(row.get("cost")),
                            "stock_value": parse_num(row.get("stock_value")),
                            "cost_value": parse_num(row.get("cost_value")),
                            "risk_score": parse_int(row.get("risk_score"), 1),
                            "decision_label": row.get("decision_label") or "",
                            "tags": row.get("tags") or "",
                            "lost_sales_value": parse_num(row.get("lost_sales_value")),
                            "frozen_capital_value": parse_num(row.get("frozen_capital_value")),
                            "stockout_probability_30d": parse_num(row.get("stockout_probability_30d")),
                        }
                    )
        except OSError:
            continue
    return rows


@app.route("/products/<path:sku>")
@login_required
def product_detail(sku: str) -> str:
    branch = selected_branch(request.args.get("branch"))
    params: list[Any] = [sku]
    branch_clause = ""
    if branch:
        branch_clause = "AND branch = ?"
        params.append(branch)
    with get_db() as db:
        current_rows = db.execute(
            f"""
            SELECT * FROM items
            WHERE sku = ? {branch_clause}
            ORDER BY risk_score DESC, lost_sales_value DESC, frozen_capital_value DESC
            """,
            params,
        ).fetchall()
        comparison_rows = db.execute(
            """
            SELECT * FROM items
            WHERE sku = ?
            ORDER BY branch
            """,
            (sku,),
        ).fetchall()

    history = product_history(sku)
    product_name = ""
    if current_rows:
        product_name = current_rows[0]["name"]
    elif history:
        product_name = history[-1]["name"]
    else:
        product_name = sku

    max_history_sales = max(max([row["sold"] for row in history] or [1]), 1)
    max_branch_stock = max(max([row["stock"] for row in comparison_rows] or [1]), 1)
    return render_template(
        "product.html",
        sku=sku,
        product_name=product_name,
        selected_branch=branch,
        branch_options=branch_options(),
        current_rows=current_rows,
        comparison_rows=comparison_rows,
        history_rows=history,
        max_history_sales=max_history_sales,
        max_branch_stock=max_branch_stock,
    )


@app.route("/financial")
@login_required
def financial() -> str:
    branch = selected_branch(request.args.get("branch"))
    return render_template(
        "financial.html",
        **financial_rows(branch),
        selected_branch=branch,
        branch_options=branch_options(),
        branch_args=branch_query_args(branch),
    )


@app.route("/weekly-comparison")
@login_required
def weekly_comparison() -> str:
    branch = selected_branch(request.args.get("branch"))
    comparison = branch_weekly_comparison(branch)
    available_branches = sorted({row["branch"] for row in branch_weekly_comparison()["rows"]})
    return render_template(
        "weekly_comparison.html",
        rows=comparison["rows"],
        totals=comparison["totals"],
        metrics=WEEKLY_COMPARISON_METRICS,
        selected_branch=branch,
        branch_options=available_branches,
        branch_args=branch_query_args(branch),
    )


@app.route("/uploads")
@login_required
def uploads() -> str:
    with get_db() as db:
        rows = db.execute("SELECT * FROM uploads ORDER BY id DESC").fetchall()
    return render_template("uploads.html", rows=rows)


@app.route("/data-management", methods=["GET", "POST"])
@login_required
def data_management() -> str | Response:
    if request.method == "POST":
        mode = request.form.get("mode", "")
        confirm = request.form.get("confirm", "").strip()
        delete_files = request.form.get("delete_files") == "1"

        if mode == "all":
            if confirm != "DELETE ALL":
                flash("Type DELETE ALL to confirm formatting all branch data.", "error")
                return redirect(url_for("data_management"))
            result = delete_branch_data(None, delete_files)
            flash(
                f"Formatted all branches: deleted {result['items']} items, {result['uploads']} imports, and {result['files']} file folders.",
                "success",
            )
            return redirect(url_for("data_management"))

        if mode == "branch":
            branch = selected_branch(request.form.get("branch"))
            if not branch:
                flash("Choose a valid branch from i-1 to i-10.", "error")
                return redirect(url_for("data_management"))
            if confirm != branch:
                flash(f"Type {branch} to confirm deleting this branch.", "error")
                return redirect(url_for("data_management"))
            result = delete_branch_data(branch, delete_files)
            flash(
                f"Deleted {branch}: removed {result['items']} items, {result['uploads']} imports, and {result['files']} file folders.",
                "success",
            )
            return redirect(url_for("data_management"))

        flash("Choose what you want to delete first.", "error")
        return redirect(url_for("data_management"))

    summaries = branch_data_summary()
    totals = {
        "items": sum(row["item_count"] for row in summaries),
        "uploads": sum(row["upload_count"] for row in summaries),
        "stock_value": sum(float(row["stock_value"] or 0) for row in summaries),
        "lost_sales_value": sum(float(row["lost_sales_value"] or 0) for row in summaries),
    }
    return render_template(
        "data_management.html",
        summaries=summaries,
        totals=totals,
        branch_choices=[f"i-{number}" for number in range(1, 11)],
    )


@app.route("/export/<category>")
@login_required
def export(category: str) -> Response:
    branch = selected_branch(request.args.get("branch"))
    rows = fetch_items(category, branch=branch)
    show_offer_columns = category in {"slow", "dead"}
    output = io.StringIO()
    writer = csv.writer(output)
    headers = [
        "Period",
        "Branch",
        "Item",
        "Ref",
        "Sales",
        "Stock",
        "Price",
    ]
    if show_offer_columns:
        headers.extend(
            [
                "Cost Price",
                "Unit Profit",
                "Gross Margin %",
                "Break Even Price",
                "Max Safe Discount %",
                "Suggested Offer Price",
                "Offer Discount %",
                "Offer Profit Per Unit",
                "Offer Profit If Sold",
            ]
        )
    headers.extend([
        "Stock Value",
        "Cost Value",
        "Avg Daily Sales",
        "Days Cover",
        "Stockout 7d",
        "Stockout 14d",
        "Stockout 30d",
        "Reorder Point",
        "Recommended Reorder Qty",
        "Risk Score",
        "Lost Sales Value",
        "Frozen Capital",
        "Signal",
        "Decision",
    ])
    writer.writerow(headers)
    for row in rows:
        values = [
            row["week_date"],
            row["branch"],
            row["name"],
            row["sku"],
            row["sold"],
            row["stock"],
            row["price"],
        ]
        if show_offer_columns:
            values.extend(
                [
                    row["cost_price"],
                    row["unit_profit"],
                    row["gross_margin_percent"],
                    row["break_even_price"],
                    row["max_safe_discount_percent"],
                    row["suggested_offer_price"],
                    row["offer_discount_percent"],
                    row["offer_profit_per_unit"],
                    row["offer_profit_if_sold"],
                ]
            )
        values.extend([
            row["stock_value"],
            row["cost_value"],
            row["avg_daily_sales"],
            row["days_cover_forecast"],
            row["stockout_probability_7d"],
            row["stockout_probability_14d"],
            row["stockout_probability_30d"],
            row["reorder_point"],
            row["recommended_reorder_qty"],
            row["risk_score"],
            row["lost_sales_value"],
            row["frozen_capital_value"],
            row["tags"],
            row["decision_label"] or row["action"],
        ])
        writer.writerow(values)
    return Response(
        "\ufeff" + output.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=pharmacy_{category}.csv"},
    )


@app.route("/login", methods=["GET", "POST"])
def login() -> str | Response:
    if session.get("user_id"):
        return redirect(url_for("index"))

    if request.method == "GET":
        return render_template("login.html")

    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")
    with get_db() as db:
        user = db.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()

    if not user or not check_password_hash(user["password_hash"], password):
        flash("اسم المستخدم أو كلمة المرور غير صحيحة.", "error")
        return redirect(url_for("login"))

    lang = current_language()
    session.clear()
    session["user_id"] = user["id"]
    session["username"] = user["username"]
    session["lang"] = lang
    return redirect(request.args.get("next") or url_for("index"))


@app.route("/logout", methods=["POST"])
def logout() -> Response:
    lang = current_language()
    session.clear()
    session["lang"] = lang
    flash("تم تسجيل الخروج بنجاح.", "success")
    return redirect(url_for("login"))


@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password() -> str | Response:
    if request.method == "GET":
        return render_template("change_password.html")

    current_password = request.form.get("current_password", "")
    new_password = request.form.get("new_password", "")
    confirm_password = request.form.get("confirm_password", "")

    if len(new_password) < 8:
        flash("كلمة المرور الجديدة يجب أن تكون 8 أحرف على الأقل.", "error")
        return redirect(url_for("change_password"))
    if new_password != confirm_password:
        flash("تأكيد كلمة المرور غير مطابق.", "error")
        return redirect(url_for("change_password"))

    with get_db() as db:
        user = db.execute("SELECT * FROM users WHERE id = ?", (session["user_id"],)).fetchone()
        if not user or not check_password_hash(user["password_hash"], current_password):
            flash("كلمة المرور الحالية غير صحيحة.", "error")
            return redirect(url_for("change_password"))
        db.execute(
            "UPDATE users SET password_hash = ? WHERE id = ?",
            (generate_password_hash(new_password), session["user_id"]),
        )

    flash("تم تغيير كلمة المرور بنجاح.", "success")
    return redirect(url_for("index"))


init_db()


if __name__ == "__main__":
    app.run(debug=False, use_reloader=False)
