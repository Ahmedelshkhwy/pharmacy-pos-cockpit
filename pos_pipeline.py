from __future__ import annotations

import csv
import json
import math
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


DEFAULT_PERIOD_DAYS = 90
DEFAULT_LEAD_TIME_DAYS = 7
TARGET_COVER_DAYS = 30
OVERSTOCK_COVER_DAYS = 180
SLOW_MOVING_RATIO = 0.05
SLOW_MOVING_COVER_DAYS = 90
MIN_OFFER_MARGIN_RATE = 0.05
DEFAULT_OFFER_DISCOUNT_RATE = 0.15

NUMERIC_COLUMNS = {
    "purchases",
    "supplier_returns",
    "adjustments",
    "b2b_sales",
    "pos_sales",
    "customer_returns",
    "internal_receives",
    "internal_sends",
    "on_hand",
    "cost",
    "sales_price",
    "tax",
    "value",
    "net_sales",
    "qoh",
    "stock_sales_value",
}

HEADER_ALIASES = {
    "ref": "ref",
    "product ref": "ref",
    "item code": "ref",
    "barcode": "barcode",
    "name": "name",
    "product name": "name",
    "item name": "name",
    "category": "category",
    "category2": "category_2",
    "category 2": "category_2",
    "category3": "category_3",
    "purchases": "purchases",
    "supplierreturns": "supplier_returns",
    "supplier returns": "supplier_returns",
    "adjustments": "adjustments",
    "b2bsales": "b2b_sales",
    "b2b sales": "b2b_sales",
    "possales": "pos_sales",
    "pos sales": "pos_sales",
    "customerreturns": "customer_returns",
    "customer returns": "customer_returns",
    "internalreceives": "internal_receives",
    "internal receives": "internal_receives",
    "internalsends": "internal_sends",
    "internal sends": "internal_sends",
    "onhand": "on_hand",
    "on hand": "on_hand",
    "cost": "cost",
    "sales_price": "sales_price",
    "sales price": "sales_price",
    "sale price": "sales_price",
    "tax": "tax",
    "value": "value",
    "netsales": "net_sales",
    "net sales": "net_sales",
    "qoh": "qoh",
    "qty": "qty",
    "total_sales_price": "stock_sales_value",
    "total sales price": "stock_sales_value",
    "cost price balance": "cost_balance",
    "sale price balance": "sale_balance",
    "branch": "branch",
}


@dataclass(frozen=True)
class PosSourceFiles:
    activity: Path
    stock: Path
    stock_cost: Path | None = None


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().casefold()
    text = re.sub(r"[_\s]+", " ", text)
    return HEADER_ALIASES.get(text, HEADER_ALIASES.get(text.replace(" ", ""), text.replace(" ", "_")))


def parse_num(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^0-9.\-]", "", str(value))
    try:
        return float(cleaned) if cleaned else 0.0
    except ValueError:
        return 0.0


def clean_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_branch_name(value: str, fallback_prefix: str = "") -> str:
    text = clean_text(value)
    match = re.search(r"(?:pharmacy|branch|i)[\s_-]*(\d{1,2})", text, flags=re.IGNORECASE)
    if match:
        number = int(match.group(1))
        if 1 <= number <= 10:
            return f"i-{number}"
    prefix_match = re.match(r"i-(\d{1,2})-?", fallback_prefix, flags=re.IGNORECASE)
    if prefix_match:
        return f"i-{int(prefix_match.group(1))}"
    return text or "i-unknown"


def parse_report_datetime(value: str) -> datetime | None:
    text = clean_text(value)
    for fmt in ("%d-%m-%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%m/%d/%Y %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def report_period_days(metadata: dict[str, str]) -> int:
    start = parse_report_datetime(metadata.get("start_date", ""))
    end = parse_report_datetime(metadata.get("end_date", ""))
    if not start or not end:
        return DEFAULT_PERIOD_DAYS
    return max((end.date() - start.date()).days + 1, 1)


def movement_period_days(metadata: dict[str, str], override_days: int | None = None) -> int:
    if override_days is not None and override_days > 0:
        return override_days
    return DEFAULT_PERIOD_DAYS


def poisson_stockout_probability(avg_daily_sales: float, qoh: float, horizon_days: int) -> float:
    if avg_daily_sales <= 0:
        return 0.0
    demand_mean = avg_daily_sales * horizon_days
    stock_units = max(math.floor(qoh), 0)
    if qoh <= 0:
        return 1.0
    if demand_mean > 80:
        z = (stock_units + 0.5 - demand_mean) / math.sqrt(demand_mean)
        cdf = 0.5 * (1 + math.erf(z / math.sqrt(2)))
        return round(max(0.0, min(1.0, 1 - cdf)), 4)

    term = math.exp(-demand_mean)
    cdf = term
    for k in range(1, stock_units + 1):
        term *= demand_mean / k
        cdf += term
    return round(max(0.0, min(1.0, 1 - cdf)), 4)


def predictive_metrics(
    *,
    pos_sales: float,
    qoh: float,
    sales_price: float,
    cost: float,
    stock_value: float,
    cost_value: float,
    turnover_ratio: float,
    period_days: int,
) -> dict[str, float | str]:
    avg_daily_sales = pos_sales / period_days if period_days > 0 else 0.0
    days_cover = qoh / avg_daily_sales if avg_daily_sales > 0 else None
    stockout_probability_7d = poisson_stockout_probability(avg_daily_sales, qoh, 7)
    stockout_probability_14d = poisson_stockout_probability(avg_daily_sales, qoh, 14)
    stockout_probability_30d = poisson_stockout_probability(avg_daily_sales, qoh, 30)

    lead_time_demand = avg_daily_sales * DEFAULT_LEAD_TIME_DAYS
    safety_stock = 1.65 * math.sqrt(lead_time_demand) if lead_time_demand > 0 else 0.0
    reorder_point = lead_time_demand + safety_stock
    target_stock = reorder_point + avg_daily_sales * TARGET_COVER_DAYS
    recommended_reorder_qty = max(math.ceil(target_stock - qoh), 0) if avg_daily_sales > 0 else 0
    expected_lost_units_30d = max(avg_daily_sales * 30 - qoh, 0)
    lost_sales_value = expected_lost_units_30d * sales_price
    unit_profit = max(sales_price - cost, 0) if sales_price > 0 and cost > 0 else 0.0
    gross_margin_percent = ((sales_price - cost) / sales_price * 100) if sales_price > 0 and cost > 0 else 0.0
    max_safe_discount_percent = max(gross_margin_percent, 0.0)
    break_even_price = cost if cost > 0 else 0.0
    target_offer_price = sales_price * (1 - DEFAULT_OFFER_DISCOUNT_RATE) if sales_price > 0 else 0.0
    minimum_profitable_price = cost * (1 + MIN_OFFER_MARGIN_RATE) if cost > 0 else 0.0
    suggested_offer_price = max(target_offer_price, minimum_profitable_price)
    if sales_price > 0:
        suggested_offer_price = min(suggested_offer_price, sales_price)
    else:
        suggested_offer_price = 0.0
    offer_discount_percent = ((sales_price - suggested_offer_price) / sales_price * 100) if sales_price > 0 else 0.0
    offer_profit_per_unit = max(suggested_offer_price - cost, 0) if cost > 0 else 0.0
    offer_profit_if_sold = offer_profit_per_unit * qoh

    frozen_capital_value = 0.0
    if qoh > 0 and (pos_sales == 0 or (days_cover is not None and days_cover > OVERSTOCK_COVER_DAYS)):
        frozen_capital_value = cost_value or stock_value

    risk_score = calculate_risk_score(
        stockout_probability_7d=stockout_probability_7d,
        stockout_probability_30d=stockout_probability_30d,
        qoh=qoh,
        pos_sales=pos_sales,
        stock_value=stock_value,
        days_cover=days_cover,
        turnover_ratio=turnover_ratio,
    )
    decision_label = decision_for_item(
        qoh=qoh,
        pos_sales=pos_sales,
        days_cover=days_cover,
        risk_score=risk_score,
        stockout_probability_7d=stockout_probability_7d,
        stockout_probability_14d=stockout_probability_14d,
        stockout_probability_30d=stockout_probability_30d,
    )
    signal_color = "red" if risk_score >= 8 else "orange" if risk_score >= 5 else "green"

    return {
        "avg_daily_sales": round(avg_daily_sales, 4),
        "days_cover": round(days_cover, 2) if days_cover is not None else None,
        "stockout_probability_7d": stockout_probability_7d,
        "stockout_probability_14d": stockout_probability_14d,
        "stockout_probability_30d": stockout_probability_30d,
        "reorder_point": round(reorder_point, 2),
        "recommended_reorder_qty": recommended_reorder_qty,
        "risk_score": risk_score,
        "decision_label": decision_label,
        "lost_sales_value": round(lost_sales_value, 2),
        "frozen_capital_value": round(frozen_capital_value, 2),
        "unit_profit": round(unit_profit, 2),
        "gross_margin_percent": round(gross_margin_percent, 2),
        "break_even_price": round(break_even_price, 2),
        "max_safe_discount_percent": round(max_safe_discount_percent, 2),
        "suggested_offer_price": round(suggested_offer_price, 2),
        "offer_discount_percent": round(max(offer_discount_percent, 0.0), 2),
        "offer_profit_per_unit": round(offer_profit_per_unit, 2),
        "offer_profit_if_sold": round(offer_profit_if_sold, 2),
        "signal_color": signal_color,
    }


def calculate_risk_score(
    *,
    stockout_probability_7d: float,
    stockout_probability_30d: float,
    qoh: float,
    pos_sales: float,
    stock_value: float,
    days_cover: float | None,
    turnover_ratio: float,
) -> int:
    score = 1.0
    score += stockout_probability_7d * 3.0
    score += stockout_probability_30d * 2.0
    if qoh <= 0 and pos_sales > 0:
        score += 3.0
    if qoh > 0 and pos_sales == 0:
        score += 2.5
    if days_cover is not None and days_cover > OVERSTOCK_COVER_DAYS:
        score += 2.0
    if turnover_ratio >= 1 and pos_sales >= 20:
        score += 1.2
    if stock_value >= 1000:
        score += 1.0
    elif stock_value >= 300:
        score += 0.5
    return int(max(1, min(10, round(score))))


def decision_for_item(
    *,
    qoh: float,
    pos_sales: float,
    days_cover: float | None,
    risk_score: int,
    stockout_probability_7d: float,
    stockout_probability_14d: float,
    stockout_probability_30d: float,
) -> str:
    if qoh <= 0 and pos_sales > 0:
        return "Buy or transfer now"
    if stockout_probability_7d >= 0.6:
        return "Buy within 7 days"
    if stockout_probability_14d >= 0.6:
        return "Plan reorder within 14 days"
    if qoh > 0 and pos_sales == 0:
        return "Clear dead stock"
    if days_cover is not None and days_cover > OVERSTOCK_COVER_DAYS:
        return "Stop buying and transfer/promote"
    if stockout_probability_30d >= 0.5:
        return "Monitor for reorder"
    if risk_score >= 8:
        return "Review today"
    return "Healthy"


def find_header_row(rows: list[tuple[Any, ...]], required_any: set[str]) -> int:
    best_index = -1
    best_score = 0
    for index, row in enumerate(rows):
        normalized = {normalize_header(cell) for cell in row if cell not in (None, "")}
        score = len(normalized & required_any)
        if score > best_score:
            best_index = index
            best_score = score
    if best_index < 0 or best_score < 2:
        raise ValueError(f"Could not detect table header. Expected any of: {sorted(required_any)}")
    return best_index


def workbook_rows(path: Path) -> list[tuple[Any, ...]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        return list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()


def extract_report_metadata(rows: list[tuple[Any, ...]]) -> dict[str, str]:
    metadata: dict[str, str] = {}
    for index, row in enumerate(rows[:12]):
        values = [clean_text(v) for v in row]
        first = values[0] if values else ""
        if first.startswith("Report Time"):
            metadata["report_time"] = first.replace("Report Time :", "").strip()
        if first == "Start Date" and index + 1 < len(rows):
            next_values = [clean_text(v) for v in rows[index + 1]]
            metadata["start_date"] = next_values[0] if len(next_values) > 0 else ""
            metadata["end_date"] = next_values[1] if len(next_values) > 1 else ""
            metadata["branch"] = next_values[2] if len(next_values) > 2 else ""
        if first == "Branch" and index + 1 < len(rows):
            next_values = [clean_text(v) for v in rows[index + 1]]
            metadata["branch"] = next_values[0] if next_values else metadata.get("branch", "")
    return metadata


def load_table(path: Path, required_any: set[str]) -> tuple[list[dict[str, Any]], dict[str, str]]:
    rows = workbook_rows(path)
    header_index = find_header_row(rows, required_any)
    headers = [normalize_header(cell) for cell in rows[header_index]]
    records: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        record = {}
        for header, value in zip(headers, row):
            if not header:
                continue
            record[header] = parse_num(value) if header in NUMERIC_COLUMNS else clean_text(value)
        if any(value not in ("", None, 0, 0.0) for value in record.values()):
            records.append(record)
    return records, extract_report_metadata(rows)


def discover_pos_files(source_dir: Path, prefix: str = "i-7-") -> PosSourceFiles:
    files = {path.name.upper(): path for path in source_dir.glob(f"{prefix}*.xlsx") if not path.name.startswith("~$")}
    activity = next((path for name, path in files.items() if "PRODUCT ACTIVITY" in name), None)
    stock = next((path for name, path in files.items() if name.endswith("STOCK.XLSX")), None)
    stock_cost = next((path for name, path in files.items() if "STOCK COST" in name), None)
    if not activity or not stock:
        raise FileNotFoundError(f"Expected {prefix}PRODUCT ACTIVITY.xlsx and {prefix}STOCK.xlsx in {source_dir}")
    return PosSourceFiles(activity=activity, stock=stock, stock_cost=stock_cost)


def index_by_ref(records: Iterable[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    indexed = {}
    for record in records:
        ref = clean_text(record.get("ref"))
        if ref:
            indexed[ref] = record
    return indexed


def merge_activity_stock(
    activity_rows: list[dict[str, Any]],
    stock_rows: list[dict[str, Any]],
    period_days: int = DEFAULT_PERIOD_DAYS,
    branch_override: str | None = None,
) -> list[dict[str, Any]]:
    stock_by_ref = index_by_ref(stock_rows)
    merged: list[dict[str, Any]] = []
    all_refs = sorted({clean_text(row.get("ref")) for row in activity_rows if row.get("ref")} | set(stock_by_ref))
    activity_by_ref = index_by_ref(activity_rows)

    for ref in all_refs:
        activity = activity_by_ref.get(ref, {})
        stock = stock_by_ref.get(ref, {})
        pos_sales = parse_num(activity.get("pos_sales"))
        qoh = parse_num(stock.get("qoh") if "qoh" in stock else activity.get("on_hand"))
        sales_price = parse_num(stock.get("sales_price") if stock.get("sales_price") else activity.get("sales_price"))
        cost = parse_num(activity.get("cost"))
        stock_value = parse_num(stock.get("stock_sales_value")) or qoh * sales_price
        cost_value = qoh * cost
        gross_profit_estimate = pos_sales * max(sales_price - cost, 0)
        turnover_ratio = pos_sales / qoh if qoh > 0 else 0
        metrics = predictive_metrics(
            pos_sales=pos_sales,
            qoh=qoh,
            sales_price=sales_price,
            cost=cost,
            stock_value=stock_value,
            cost_value=cost_value,
            turnover_ratio=turnover_ratio,
            period_days=period_days,
        )
        days_cover = metrics["days_cover"]
        tags = classify_item(pos_sales=pos_sales, qoh=qoh, turnover_ratio=turnover_ratio, days_cover=days_cover)
        item_branch = branch_override or normalize_branch_name(clean_text(stock.get("branch") or activity.get("branch")))
        item = {
                "ref": ref,
                "barcode": clean_text(activity.get("barcode")),
                "name": clean_text(activity.get("name") or stock.get("name")),
                "category": clean_text(activity.get("category") or "UNCATEGORIZED"),
                "branch": item_branch,
                "purchases": parse_num(activity.get("purchases")),
                "pos_sales": pos_sales,
                "net_sales": parse_num(activity.get("net_sales")),
                "qoh": qoh,
                "sales_price": sales_price,
                "cost": cost,
                "stock_value": stock_value,
                "cost_value": cost_value,
                "gross_profit_estimate": gross_profit_estimate,
                "turnover_ratio": turnover_ratio,
                "days_cover": days_cover,
                "tags": " | ".join(tags) if tags else "normal",
        }
        item.update(metrics)
        merged.append(item)
    return merged


def classify_item(pos_sales: float, qoh: float, turnover_ratio: float, days_cover: float | None = None) -> list[str]:
    tags: list[str] = []
    if qoh > 0 and pos_sales == 0:
        tags.append("dead_stock")
    if qoh > 0 and pos_sales > 0 and (
        turnover_ratio < SLOW_MOVING_RATIO
        or (days_cover is not None and days_cover > SLOW_MOVING_COVER_DAYS)
    ):
        tags.append("slow_moving")
    if qoh <= 0 and pos_sales > 0:
        tags.append("stockout_with_sales")
    if pos_sales >= 20 and turnover_ratio >= 1:
        tags.append("fast_moving")
    if qoh > 0 and pos_sales > 0 and days_cover is not None and days_cover > OVERSTOCK_COVER_DAYS:
        tags.append("overstock_risk")
    return tags


def summarize(items: list[dict[str, Any]], metadata: dict[str, str]) -> dict[str, Any]:
    tag_counter = Counter()
    category_summary: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for item in items:
        for tag in str(item["tags"]).split(" | "):
            tag_counter[tag] += 1
        category = item["category"] or "UNCATEGORIZED"
        category_summary[category]["items"] += 1
        category_summary[category]["pos_sales"] += item["pos_sales"]
        category_summary[category]["net_sales"] += item["net_sales"]
        category_summary[category]["stock_value"] += item["stock_value"]
        category_summary[category]["gross_profit_estimate"] += item["gross_profit_estimate"]

    return {
        "metadata": metadata,
        "total_items": len(items),
        "total_pos_sales_qty": sum(item["pos_sales"] for item in items),
        "total_net_sales": sum(item["net_sales"] for item in items),
        "total_stock_value": sum(item["stock_value"] for item in items),
        "total_cost_value": sum(item["cost_value"] for item in items),
        "gross_profit_estimate": sum(item["gross_profit_estimate"] for item in items),
        "lost_sales_value": sum(item["lost_sales_value"] for item in items),
        "frozen_capital_value": sum(item["frozen_capital_value"] for item in items),
        "high_risk_count": sum(item["risk_score"] >= 8 for item in items),
        "reorder_qty_total": sum(item["recommended_reorder_qty"] for item in items),
        "tag_counts": dict(tag_counter),
        "category_summary": dict(category_summary),
    }


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str] | None = None) -> None:
    if not rows:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    columns = fieldnames or list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def top_rows(items: list[dict[str, Any]], key: str, limit: int = 20) -> list[dict[str, Any]]:
    return sorted(items, key=lambda item: item.get(key) or 0, reverse=True)[:limit]


def filter_tag(items: list[dict[str, Any]], tag: str) -> list[dict[str, Any]]:
    return [item for item in items if tag in item["tags"]]


def build_markdown(summary: dict[str, Any], output_dir: Path) -> str:
    metadata = summary["metadata"]
    tag_counts = summary["tag_counts"]
    lines = [
        "# POS Kick-off Analysis",
        "",
        f"- Branch: {metadata.get('branch', 'Unknown')}",
        f"- Period: {metadata.get('start_date', 'Unknown')} to {metadata.get('end_date', 'Unknown')}",
        f"- Export period days: {metadata.get('actual_export_period_days', 'Unknown')}",
        f"- Movement calculation days: {metadata.get('period_days', DEFAULT_PERIOD_DAYS)}",
        f"- Report time: {metadata.get('report_time', 'Unknown')}",
        f"- Total items: {summary['total_items']:,.0f}",
        f"- Total POS sales qty: {summary['total_pos_sales_qty']:,.2f}",
        f"- Total net sales: {summary['total_net_sales']:,.2f}",
        f"- Total stock sales value: {summary['total_stock_value']:,.2f}",
        f"- Estimated stock cost value: {summary['total_cost_value']:,.2f}",
        f"- Estimated gross profit from POS sales: {summary['gross_profit_estimate']:,.2f}",
        f"- Expected lost sales next 30 days: {summary['lost_sales_value']:,.2f}",
        f"- Frozen capital value: {summary['frozen_capital_value']:,.2f}",
        f"- High risk items: {summary['high_risk_count']:,.0f}",
        f"- Recommended reorder units: {summary['reorder_qty_total']:,.0f}",
    ]
    if metadata.get("period_warning"):
        lines.extend(["", "## Period Warning", "", f"- {metadata['period_warning']}"])
    lines.extend(["", "## Signals", ""])
    for key in ["dead_stock", "slow_moving", "overstock_risk", "stockout_with_sales", "fast_moving", "normal"]:
        lines.append(f"- {key}: {tag_counts.get(key, 0):,.0f}")
    lines.extend(
        [
            "",
            "## Output Files",
            "",
            f"- merged_items.csv: {output_dir / 'merged_items.csv'}",
            f"- category_summary.csv: {output_dir / 'category_summary.csv'}",
            f"- dead_stock.csv: {output_dir / 'dead_stock.csv'}",
            f"- slow_moving.csv: {output_dir / 'slow_moving.csv'}",
            f"- stockout_with_sales.csv: {output_dir / 'stockout_with_sales.csv'}",
            f"- top_stock_value.csv: {output_dir / 'top_stock_value.csv'}",
        ]
    )
    return "\n".join(lines) + "\n"


def analyze_pos_exports(
    source_dir: Path,
    output_root: Path,
    prefix: str = "i-7-",
    branch_override: str | None = None,
    movement_period_days_override: int | None = None,
) -> Path:
    source_files = discover_pos_files(source_dir, prefix=prefix)
    activity_rows, activity_metadata = load_table(source_files.activity, {"ref", "name", "pos_sales", "on_hand"})
    stock_rows, stock_metadata = load_table(source_files.stock, {"ref", "name", "qoh", "sales_price"})
    metadata = {**activity_metadata, **{k: v for k, v in stock_metadata.items() if v and k not in activity_metadata}}
    normalized_override = normalize_branch_name(branch_override or "", fallback_prefix=prefix) if branch_override else None
    metadata["branch"] = normalized_override or normalize_branch_name(metadata.get("branch", ""), fallback_prefix=prefix)
    actual_period_days = report_period_days(metadata)
    period_days = movement_period_days(metadata, movement_period_days_override)
    if movement_period_days_override is not None:
        if not metadata.get("start_date") or not metadata.get("end_date"):
            raise ValueError("Could not read Product Activity Start Date and End Date; the selected analysis period cannot be verified.")
        if actual_period_days != period_days:
            raise ValueError(
                f"Product Activity export covers {actual_period_days} days, "
                f"but the selected analysis period is {period_days} days."
            )
    metadata["actual_export_period_days"] = str(actual_period_days)
    metadata["period_days"] = str(period_days)
    metadata["selected_period_days"] = str(period_days)
    if actual_period_days != period_days:
        metadata["period_warning"] = (
            f"Product Activity export covers {actual_period_days} days, "
            f"while movement calculations use {period_days} days. "
            f"For best accuracy, export Product Activity for exactly {period_days} days."
        )
    items = merge_activity_stock(activity_rows, stock_rows, period_days=period_days, branch_override=normalized_override)
    summary = summarize(items, metadata)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = output_root / f"{prefix.strip('-')}_{stamp}"
    output_dir.mkdir(parents=True, exist_ok=True)

    item_columns = [
        "ref",
        "barcode",
        "name",
        "category",
        "branch",
        "purchases",
        "pos_sales",
        "net_sales",
        "qoh",
        "sales_price",
        "cost",
        "stock_value",
        "cost_value",
        "gross_profit_estimate",
        "turnover_ratio",
        "days_cover",
        "avg_daily_sales",
        "stockout_probability_7d",
        "stockout_probability_14d",
        "stockout_probability_30d",
        "reorder_point",
        "recommended_reorder_qty",
        "risk_score",
        "decision_label",
        "lost_sales_value",
        "frozen_capital_value",
        "unit_profit",
        "gross_margin_percent",
        "break_even_price",
        "max_safe_discount_percent",
        "suggested_offer_price",
        "offer_discount_percent",
        "offer_profit_per_unit",
        "offer_profit_if_sold",
        "signal_color",
        "tags",
    ]
    write_csv(output_dir / "merged_items.csv", items, item_columns)
    category_rows = [
        {"category": category, **values}
        for category, values in sorted(
            summary["category_summary"].items(),
            key=lambda entry: entry[1]["net_sales"],
            reverse=True,
        )
    ]
    write_csv(output_dir / "category_summary.csv", category_rows)
    write_csv(output_dir / "dead_stock.csv", top_rows(filter_tag(items, "dead_stock"), "stock_value", 200), item_columns)
    write_csv(output_dir / "slow_moving.csv", top_rows(filter_tag(items, "slow_moving"), "stock_value", 200), item_columns)
    write_csv(output_dir / "stockout_with_sales.csv", top_rows(filter_tag(items, "stockout_with_sales"), "pos_sales", 200), item_columns)
    write_csv(output_dir / "top_stock_value.csv", top_rows(items, "stock_value", 200), item_columns)
    write_csv(output_dir / "top_net_sales.csv", top_rows(items, "net_sales", 200), item_columns)

    (output_dir / "summary.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    (output_dir / "README.md").write_text(build_markdown(summary, output_dir), encoding="utf-8")
    return output_dir
