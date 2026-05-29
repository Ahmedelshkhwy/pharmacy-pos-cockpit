from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_PATH = BASE_DIR / "pharmacy_branch_upload_template.xlsx"

HEADERS = [
    "اسم الصنف",
    "كود الصنف",
    "الكمية المباعة",
    "الرصيد الحالي",
    "تاريخ انتهاء الصلاحية",
    "سعر الوحدة",
]

EXAMPLE_ROWS = [
    ["Panadol Extra 24 Tabs", "RX-1001", 0, 85, date.today() + timedelta(days=18), 12.5],
    ["Augmentin 1g Tablets", "RX-1002", 6, 44, date.today() + timedelta(days=36), 72],
    ["Vitamin D3 50000 IU", "RX-1003", 2, 120, date.today() + timedelta(days=80), 38],
    ["Cetirizine 10mg", "RX-1004", 76, 38, date.today() + timedelta(days=210), 9.5],
]


def style_header(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="16324F")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def build_workbook() -> Workbook:
    wb = Workbook()
    data_sheet = wb.active
    data_sheet.title = "بيانات الفرع"
    data_sheet.sheet_view.rightToLeft = True
    data_sheet.append(HEADERS)
    for row in EXAMPLE_ROWS:
        data_sheet.append(row)

    style_header(data_sheet)
    data_sheet.freeze_panes = "A2"
    data_sheet.auto_filter.ref = "A1:F200"

    widths = [30, 18, 18, 18, 22, 16]
    for index, width in enumerate(widths, start=1):
        data_sheet.column_dimensions[chr(64 + index)].width = width

    for row in data_sheet.iter_rows(min_row=2, max_row=200, min_col=3, max_col=4):
        for cell in row:
            cell.number_format = "0"
    for cell in data_sheet["E"][1:200]:
        cell.number_format = "yyyy-mm-dd"
    for cell in data_sheet["F"][1:200]:
        cell.number_format = "0.00"

    whole_number = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    decimal_number = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    data_sheet.add_data_validation(whole_number)
    data_sheet.add_data_validation(decimal_number)
    whole_number.add("C2:D200")
    decimal_number.add("F2:F200")

    instructions = wb.create_sheet("تعليمات")
    instructions.sheet_view.rightToLeft = True
    instructions["A1"] = "نموذج رفع تقرير الفرع الأسبوعي"
    instructions["A1"].font = Font(size=16, bold=True, color="16324F")
    instructions["A3"] = "املأ شيت بيانات الفرع فقط، ثم ارفع الملف من صفحة رفع تقرير داخل المنصة."
    instructions["A5"] = "اسم الملف المقترح:"
    instructions["B5"] = "اسم_الفرع_YYYY-MM-DD.xlsx"
    instructions["A7"] = "الأعمدة المطلوبة:"
    instructions["A8"] = "اسم الصنف"
    instructions["B8"] = "اسم المنتج كما يظهر في النظام"
    instructions["A9"] = "كود الصنف"
    instructions["B9"] = "SKU أو Barcode"
    instructions["A10"] = "الكمية المباعة"
    instructions["B10"] = "إجمالي مبيعات الأسبوع"
    instructions["A11"] = "الرصيد الحالي"
    instructions["B11"] = "كمية المخزون وقت التصدير"
    instructions["A12"] = "تاريخ انتهاء الصلاحية"
    instructions["B12"] = "يفضل yyyy-mm-dd"
    instructions["A13"] = "سعر الوحدة"
    instructions["B13"] = "سعر بيع/تكلفة الوحدة حسب المتاح"
    instructions["A15"] = "مهم:"
    instructions["B15"] = "لا تغير أسماء الأعمدة في الصف الأول."

    for row in instructions.iter_rows(min_row=5, max_row=15, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    instructions.column_dimensions["A"].width = 28
    instructions.column_dimensions["B"].width = 58

    return wb


if __name__ == "__main__":
    workbook = build_workbook()
    workbook.save(OUTPUT_PATH)
    print(OUTPUT_PATH)
